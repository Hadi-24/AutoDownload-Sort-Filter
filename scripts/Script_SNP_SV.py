import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define file paths
file_path = "/Users/"  # Main Excel file,this is the excel with all the data that will be read (change path)
svs_file_path = "/Users/"  # Separate output excel for SVs (change path)
sheet_name = ""  # The name of the sheet you want to process

# Read the data from the Excel file (in this case the sheet "Raw")
data = pd.read_excel(file_path, sheet_name=sheet_name)

def one_char_difference(spdi):
    """Check if exactly one character difference occurred in Canonical SPDI."""
    if pd.isna(spdi) or not isinstance(spdi, str) or ":" not in spdi:
        return False
    
    parts = spdi.split(":")
    if len(parts) < 4:
        return False
    
    ref = parts[2]  # Reference allele
    alt = parts[3]  # Alternate allele
    
    # If Ref and Alt are exactly the same, it's not a SNP
    if ref == alt:
        return False
    
    # Fix for single-character deletions or insertions
    if len(ref) == 1 and len(alt) == 0:
        return True
    if len(ref) == 0 and len(alt) == 1:
        return True

    # Prevent multi-character deletions/insertions from being classified as SNPs
    if len(ref) > 1 and len(alt) == 0:
        return False
    if len(ref) == 0 and len(alt) > 1:
        return False

    # Count differences
    ref_chars = list(ref)
    alt_chars = list(alt)
    min_length = min(len(ref_chars), len(alt_chars))
    diff_count = sum(1 for i in range(min_length) if ref_chars[i] != alt_chars[i])
    diff_count += abs(len(ref_chars) - len(alt_chars))

    return diff_count == 1

# Extract Ref and Alt from the 'Canonical SPDI' column
def extract_ref(spdi):
    if isinstance(spdi, str):
        parts = spdi.split(":")
        return parts[2] if len(parts) >= 3 else None
    return None

def extract_alt(spdi):
    if isinstance(spdi, str):
        parts = spdi.split(":")
        return parts[3] if len(parts) >= 4 else None
    return None

# Extract Start and End positions from GRCh38Location using regex
data['Start'] = data['GRCh38Location'].astype(str).str.extract(r'^(\d+)', expand=False)
data['End'] = data['GRCh38Location'].astype(str).str.extract(r'(\d+)$', expand=False)

# Extract Ref and Alt
data['Ref'] = data['Canonical SPDI'].apply(extract_ref)
data['Alt'] = data['Canonical SPDI'].apply(extract_alt)

# Filter SNPs based on the one_char_difference function
filtered_snps = data[data['Canonical SPDI'].apply(one_char_difference)].copy()

# Create SNPs_Final with only Start, End, Ref, Alt
snps_final = filtered_snps[['Start', 'End', 'Ref', 'Alt']].copy()

# Create Filtered_SNPs without Start, End, Ref, Alt
filtered_snps = filtered_snps.drop(columns=['Start', 'End', 'Ref', 'Alt'])

# Create SVs: rows that were filtered out
svs = data[~data['Canonical SPDI'].isin(filtered_snps['Canonical SPDI'])].copy()

# Select only the necessary columns for SVs_Final
svs_final = svs[['Start', 'End', 'Ref', 'Alt']].copy()

# Drop Start, End, Ref, and Alt from the SVs sheet
svs = svs.drop(columns=['Start', 'End', 'Ref', 'Alt'])

# Save SNP Data in the Original File (you can save it to separate file if you want I just saved it to the same ine)

wb = load_workbook(file_path)

# Define new sheet names (change them per you preference)
filtered_snps_sheet = "Filtered_SNPs"
snps_final_sheet = "SNPs_Final"

# Remove old sheets if they exist to avoid duplicates (basically overwrite old sheets if script is run again in case of changes to script)
for sheet in [filtered_snps_sheet, snps_final_sheet]:
    if sheet in wb.sheetnames:
        std = wb[sheet]
        wb.remove(std)

# Add a new sheet for Filtered_SNPs (this sheet is WITHOUT Start, End, Ref, Alt)
ws_filtered = wb.create_sheet(filtered_snps_sheet)
for row in dataframe_to_rows(filtered_snps, index=False, header=True):
    ws_filtered.append(row)

# Add a new sheet for SNPs_Final (this one ONLY contains Start, End, Ref, Alt)
ws_final = wb.create_sheet(snps_final_sheet)
for row in dataframe_to_rows(snps_final, index=False, header=True):
    ws_final.append(row)

# Save the updated SNPs file
wb.save(file_path)

# Save SV Data in a Separate Excel File

wb_svs = Workbook()

# Remove the default empty sheet
wb_svs.remove(wb_svs.active)

# Add SVs sheet (this sheet is WITHOUT Start, End, Ref, Alt)
ws_svs = wb_svs.create_sheet("SVs")
for row in dataframe_to_rows(svs, index=False, header=True):
    ws_svs.append(row)

# Add SVs_Final sheet (this sheet contains ONLY the columns Start, End, Ref, Alt)
ws_svs_final = wb_svs.create_sheet("SVs_Final")
for row in dataframe_to_rows(svs_final, index=False, header=True):
    ws_svs_final.append(row)

# Save the SVs file
wb_svs.save(svs_file_path)

# Print a single success message
print("✅ Processing successfully completed. SNPs saved in the original file, SVs saved in a separate file.")
